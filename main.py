import tkinter as tk
from tkinter import filedialog, messagebox, colorchooser, ttk
from PIL import Image, ImageDraw, ImageFont
import openpyxl
import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options


def generate_certificate(template_path, name, serial_number, output_path, font_name="arial.ttf", font_size=48, text_color="black", y_offset=0):
    img = Image.open(template_path)
    font = ImageFont.truetype(font_name, font_size)
    draw = ImageDraw.Draw(img)
    image_width, image_height = img.size

   
    serial_font_size = int(font_size * 0.6) 
    serial_font = ImageFont.truetype(font_name, serial_font_size)
    serial_text = f"S.No.: {serial_number}"
    serial_bbox = draw.textbbox((0, 0), serial_text, font=serial_font)
    serial_text_width = serial_bbox[2] - serial_bbox[0]
    serial_text_height = serial_bbox[3] - serial_bbox[1]
    serial_text_x = 80  
    serial_text_y = 20 
    draw.text((serial_text_x, serial_text_y), serial_text, font=serial_font, fill="#fff")

   
    bbox = draw.textbbox((0, 0), name, font=font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    text_x = (image_width - text_width) // 2
    text_y = (image_height - text_height) // 2 + y_offset
    draw.text((text_x, text_y), name, font=font, fill=text_color)

    img.save(output_path)
    print(f"Certificate generated for {name} with Serial Number {serial_number} at {output_path}")


def send_certificate_via_whatsapp(phone_number, certificate_path, driver):
    message = "Congratulations! Here is your certificate of participation."
    
    whatsapp_url = f"https://web.whatsapp.com/send?phone=+91{phone_number}&text={message}"
    driver.get(whatsapp_url)
    time.sleep(12)  

    attach_button = driver.find_element(By.XPATH, '//div[@title="Attach"]')
    attach_button.click()
    time.sleep(1)

    image_input = driver.find_element(By.CSS_SELECTOR, 'input[type="file"]')
    image_input.send_keys(os.path.abspath(certificate_path))

    time.sleep(4)

    send_button = driver.find_element(By.XPATH, '//span[@data-icon="send"]')
    send_button.click()

    print(f"Certificate sent to {phone_number}")
    time.sleep(4)


def process_certificates(template_path, excel_file, progress_var, y_offset=0, font_name="arial.ttf", font_size=48, text_color="black"):
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)

    service = Service("chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=chrome_options)

    driver.get("https://web.whatsapp.com")
    input("Please scan the QR code and press Enter to continue...")

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    if not os.path.exists("certificates"):
        os.makedirs("certificates")

    total_rows = sheet.max_row - 1
    processed_count = 0

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        name, phone_number = row
        if name and phone_number:
            serial_number = f"EC2024TT{str(idx).zfill(4)}" 
            output_path = f"certificates/{name}.png"
            generate_certificate(template_path, name, serial_number, output_path, font_name, font_size, text_color, y_offset)
            
            send_certificate_via_whatsapp(phone_number, output_path, driver)

            processed_count += 1
            progress_var.set((processed_count / total_rows) * 100)
            app.update_idletasks()

    print("Certificates have been sent successfully.")


def browse_template():
    filename = filedialog.askopenfilename(title="Select Certificate Template", filetypes=[("PNG Files", "*.png")])
    if filename:
        template_entry.delete(0, tk.END)
        template_entry.insert(0, filename)


def browse_excel():
    filename = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
    if filename:
        excel_entry.delete(0, tk.END)
        excel_entry.insert(0, filename)


def choose_color():
    color = colorchooser.askcolor()[1]
    if color:
        color_entry.delete(0, tk.END)
        color_entry.insert(0, color)


def start_process():
    template_path = template_entry.get()
    excel_file = excel_entry.get()
    font_name = font_entry.get()
    font_size = int(font_size_entry.get())
    text_color = color_entry.get()
    y_offset = int(y_offset_entry.get())

    if not template_path or not excel_file:
        messagebox.showerror("Error", "Please select both template and Excel file!")
        return

    try:
        process_certificates(template_path, excel_file, progress_var, y_offset, font_name, font_size, text_color)
        messagebox.showinfo("Success", "Certificates generated and sent successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


app = tk.Tk()
app.title("Certificate Generator")
app.geometry("1560x900")
app.configure(bg="#f0f0f0")


title_label = tk.Label(app, text="Certificate Generator", font=("Times", 40, "bold"), bg="#000000", fg="#ff0000", pady=20)
title_label.pack(fill=tk.X)

tk.Label(app, text="Select Certificate Template:", font=("Times", 12), bg="#f0f0f0").pack(pady=5)
template_entry = tk.Entry(app, width=50)
template_entry.pack(pady=5)
tk.Button(app, text="Browse", command=browse_template, bg="#2196F3", fg="white").pack(pady=5)


tk.Label(app, text="Select Excel File:", font=("Arial", 12), bg="#f0f0f0").pack(pady=5)
excel_entry = tk.Entry(app, width=50)
excel_entry.pack(pady=5)
tk.Button(app, text="Browse", command=browse_excel, bg="#2196F3", fg="white").pack(pady=5)


tk.Label(app, text="Font Name:", font=("Arial", 12), bg="#f0f0f0").pack(pady=5)
font_entry = tk.Entry(app, width=50)
font_entry.insert(0, "arial.ttf")
font_entry.pack(pady=5)


tk.Label(app, text="Font Size:", font=("Arial", 12), bg="#f0f0f0").pack(pady=5)
font_size_entry = tk.Entry(app, width=50)
font_size_entry.insert(0, "48")
font_size_entry.pack(pady=5)


tk.Label(app, text="Text Color:", font=("Arial", 12), bg="#f0f0f0").pack(pady=5)
color_entry = tk.Entry(app, width=50)
color_entry.insert(0, "black")
color_entry.pack(pady=5)
tk.Button(app, text="Choose Color", command=choose_color, bg="#2196F3", fg="white").pack(pady=5)


tk.Label(app, text="Y-Offset:", font=("Arial", 12), bg="#f0f0f0").pack(pady=5)
y_offset_entry = tk.Entry(app, width=50)
y_offset_entry.insert(0, "0")
y_offset_entry.pack(pady=5)


progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(app, orient="horizontal", length=400, mode="determinate", variable=progress_var)
progress_bar.pack(pady=20)


tk.Button(app, text="Generate Certificates", command=start_process, bg="#4CAF50", fg="white", font=("Arial", 14)).pack(pady=20)


developer_label = tk.Label(app, text="Developed by Devesh Rawat , Anuj Anthwal & Akhilesh Raje", font=("Times New Roman", 18), bg="#f0f0f0", fg="#4b4f50")

developer_label.pack(side=tk.BOTTOM, pady=40)


app.mainloop()
